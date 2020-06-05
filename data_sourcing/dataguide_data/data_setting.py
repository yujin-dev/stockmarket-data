import os
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import pickle


class SortStockData:

    """ DataGuide에서 엑셀로 불러온 데이터 정리 """

    def __init__(self, data, specify_folder=None):

        self.data = data

        if specify_folder is None:
            self.save_dir = os.path.join(os.getcwd(), "data_setting/sort_stock")

        else:
            self.save_dir = os.path.join(os.getcwd(), "data_setting/" + specify_folder)

        os.makedirs(self.save_dir, exist_ok=True)
        self.new_data = {}

    def read_pkl(self, data_name):

        """pickle 파일 읽기

        Parameters
        ----------------------------
        data_name: pickle 데이터 파일 이름

        Returns
        ----------------------------
        읽은 pickle 파일
        """

        df = pd.read_pickle(self.save_dir, data_name)
        return df

    def to_pkl(self, data, data_name):

        """pickle 파일 읽기

        Parameters
        ----------------------------
        data: pickle로 저장될 데이터
        data_name: 저장될 데이터 이름

        """
        with open(os.path.join(self.save_dir, data_name + ".pickle"), "wb") as fw:
            pickle.dump(data, fw)

    def sort_stock_data(self, sort_type, save_type="csv"):

        """ 데이터 가이드에서 주식 데이터 (Calendar 주기) 읽고 sort_type별로 분류해서 저장

        Parameters
        ----------------------------
        sort_type: 종목별로 파일을 생성하면 'stock', item별로 파일을 생성하면 'item'
        save_type: 저장할 형식이 csv파일이면 'csv', pickle파일이면 'pickle'
        """

        if ".csv" in self.data:

            data = pd.read_csv(
                self.data,
                index_col=0,
                header=8,
                thousands=",",
                skiprows=[9, 10, 11, 13],
                engine="python",
            )

            if sort_type == "stock":
                self.sort_stock(data)

            elif sort_type == "item":
                self.sort_item(data)

        else:  # data가 엑셀 파일인 경우( .xlsx, .xlsm )

            sheets = load_workbook(self.data).sheetnames

            for sheet in sheets:
                print(sheet)

                data = pd.read_excel(
                    self.data,
                    sheet_name=sheet,
                    index_col=0,
                    header=8,
                    thousands=",",
                    skiprows=[9, 10, 11, 13],
                )

                if sort_type == "stock":
                    self.sort_stock(data)

                elif sort_type == "item":
                    self.sort_item(data)

        for key, value in self.new_data.items():

            try:
                value.index = value.index.map(lambda x: x.strftime("%Y-%m-%d"))
            except:
                pass

            if save_type == "csv":
                print(key, value.shape)
                value.to_csv(
                    os.path.join(self.save_dir, key + ".csv"), encoding="cp949"
                )

            elif save_type == "pickle":
                print(key, value.shape)
                self.to_pkl(value, key)

            elif save_type is None:
                return 
                

    def sort_item(self, data):

        """ data를 읽고 item별로 파일을 저장( 파일 이름 : item이름 + 확장자 )
        """

        data.columns = data.columns.map(lambda x: x.split(".")[0])
        sort_data = data.T.reset_index(drop=False).set_index("Item Name").T

        for item_name in set(sort_data.columns):

            item_data = sort_data[item_name]
            item_data = item_data.T.set_index("index").T
            item_data.index.name = "date"
            item_data.columns.name = None

            if item_name != "시장구분":
                item_data = item_data.apply(pd.to_numeric)

            if "순매수대금(개인)" in item_name:
                item_name = "개인순매수"
            elif "순매수대금(기관계)" in item_name:
                item_name = "기관순매수"
            elif "순매수대금(등록외국인)" in item_name:
                item_name = "등록외국인순매수"
            elif "순매수대금(외국인계)" in item_name:
                item_name = "외인순매수"

            try:
                concat_data = self.new_data[item_name]
                concat_data = pd.concat([concat_data, item_data], axis=1)
                self.new_data.update({item_name: concat_data})
            except:
                self.new_data[item_name] = item_data


    def sort_stock(self, data):

        """ data를 읽고 종목별로 파일을 저장( 파일 이름 : 종목코드 + 확장자 )
        """

        data_column = data.columns.map(lambda x: x.split(".")[0])
        data.columns = data_column

        for stock_name in sorted(set(data_column)):
            stock_data = data[stock_name].T.set_index("Item Name").T
            stock_data.index.name = "date"
            stock_data.columns.name = None

            if stock_name != "시장구분":
                stock_data = stock_data.apply(pd.to_numeric)

            try:
                concat_data = self.new_data[stock_name]
                concat_data = pd.concat([concat_data, stock_data], axis=1)
                self.new_data.update({stock_name: concat_data})
            except:
                self.new_data[stock_name] = stock_data
